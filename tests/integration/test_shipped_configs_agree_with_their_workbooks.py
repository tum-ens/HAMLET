"""#214 -- a scenario's `agents.yaml` and its `agents.xlsx` must not contradict each other.

Two files in every workbook-bearing scenario state the modelling backend, and **which one wins
depends on the Creator entry point the scenario's own notebook calls**:

* `new_scenario_from_configs` and `new_scenario_from_grids` regenerate `agents.xlsx` from the YAML
  (`__create_agent_files` passes `overwrite=True`), so the YAML wins and a stale workbook is inert.
* `new_scenario_from_files` builds the agents *from the workbook* and never regenerates it, so the
  **workbook** wins and the YAML is decoration.

`create_scenario_with_market` was the second kind: `agents.yaml` said `poi`, `agents.xlsx` said
`linopy`, and a user who read the YAML and ran the notebook got linopy. No test ran that example,
so nothing noticed. `scenario_with_grid` had the same drift on both keys, inert only by accident
of its entry point.

**This test does not care which entry point a scenario uses, and that is the point.** Requiring
agreement means the reader gets the same answer from either file, so the entry point stops being
load-bearing information a reader has to know before trusting what they are looking at. It is also
the cheap half: it needs no run, so it lives in the fast tier and fails in milliseconds where the
consequence would otherwise surface only in a job costing minutes.

Deliberately *not* an assertion about which backend is correct. `scenario_with_market` ships
`solver: gurobi` in both files and that is a real choice, not a defect. Only disagreement is one.
"""
import re
from pathlib import Path

import pandas as pd
import pytest

REPO_ROOT = Path(__file__).resolve().parents[2]

#: The keys that appear in both files and that a mismatch on changes what runs.
#:
#: Guarded by `test_the_compared_keys_are_the_ones_that_decide_the_backend`. An allowlist nothing
#: compares against the tree exempts whatever is left out: dropping `'solver'` here re-opens #214
#: on the very file this module was written to repair, and every test still passes. Demonstrated.
#:
#: `time_limit` is in the list because that guard put it there -- it is written to the workbook
#: alongside the other two and was not being compared. It is not decoration: a `time_limit` the
#: Executor reads in seconds, where the config author meant minutes, is #204.
KEYS = ('framework', 'solver', 'time_limit')

#: Every scenario config folder in the repository, named rather than globbed.
#:
#: A glob would make this test pass by finding nothing -- rename `agents.xlsx` and the suite goes
#: green over a workbook that no longer agrees with anything. `test_every_shipped_scenario_is_
#: listed_here` below walks the tree and fails if this list has drifted from it, so the list is
#: checked rather than trusted; adding a scenario means adding it here, visibly, in the same commit.
SCENARIOS = (
    'examples/create_scenario_with_grid/scenario_with_grid',
    'examples/create_scenario_with_market/scenario_with_market',
    'examples/create_scenario_with_topology/scenario_with_topology',
    'examples/create_simple_scenario/simple_scenario',
    'tests/e2e/scenarios/ctsp_industry',
    'tests/e2e/scenarios/grid_golden',
)

#: The one scenario that ships no workbook -- the Creator writes it from the YAML. Named, so that
#: a scenario silently losing its workbook fails rather than becoming exempt.
NO_WORKBOOK = ('examples/create_simple_scenario/simple_scenario',)


def canonical(value):
    """One spelling for a value that YAML gives as text and the workbook as a number.

    `time_limit: 120` is the string `'120'` out of the YAML and the int `120` out of pandas, and
    comparing those raw reports a disagreement that is not one. Numbers compare as numbers,
    everything else as text -- so `poi` still only equals `poi`, and `120` equals `120.0` because
    the Creator writes whichever the sheet's dtype gives it.
    """
    try:
        return float(value)
    except (TypeError, ValueError):
        return str(value).strip()


def yaml_values(path, key):
    """Every value the YAML gives `key`, as a set. The Creator reads all of them, not the first."""
    pattern = re.compile(r'^\s*' + re.escape(key) + r': *([\w.]+)', re.MULTILINE)
    return {canonical(value) for value in pattern.findall(path.read_text(encoding='utf-8'))}


def workbook_values(path, key):
    """Every `.../<key>` value per sheet, read the way `create_agents_from_file` reads it.

    Reported **per sheet**, never as one set over the workbook. `create_agents_file_from_config`
    writes one sheet per agent type, so a workbook-wide answer lets a matching `sfh` sheet vouch
    for an `industry` sheet that disagrees -- which is this very defect one level down, and is a
    live risk now that `ctsp_industry` ships two sheets.

    The `ExcelFile` is closed explicitly: pandas defines no `__del__`, so an unclosed handle keeps
    the file open on Windows.
    """
    per_sheet = {}
    with pd.ExcelFile(path) as book:
        for sheet in book.sheet_names:
            frame = book.parse(sheet, index_col=0)
            values = set()
            for column in frame.columns:
                if str(column).rsplit('/', 1)[-1] == key:
                    values.update(canonical(value) for value in frame[column].dropna())
            per_sheet[sheet] = values
    return per_sheet


def mismatches(yaml_path, book_path, name, keys=KEYS):
    """Every (key, sheet) at which the workbook does not say what the YAML says.

    Returned as a list rather than raised, so the caller reports all of them at once: when two
    files drift it is the *shape* of the drift across keys and sheets that says whether it was one
    careless edit or a stale workbook. A helper rather than an inline loop so that the two stub
    tests below can drive it with workbooks no shipped scenario can produce -- without them the
    per-sheet split is untested, which it was.
    """
    found = []
    for key in keys:
        expected = yaml_values(yaml_path, key)
        for sheet, values in workbook_values(book_path, key).items():
            if values != expected:
                found.append(f'{name}: agents.yaml says {key}={sorted(expected)} but agents.xlsx '
                             f'sheet {sheet!r} says {sorted(values)}')
    return found


@pytest.fixture(params=SCENARIOS, ids=lambda path: path.rsplit('/', 1)[-1])
def scenario(request):
    """One scenario config folder, as an absolute path."""
    return REPO_ROOT / request.param, request.param


def test_every_shipped_scenario_is_listed_here():
    """`SCENARIOS` is the tree, not a subset of it that happens to pass.

    Without this the list is an allowlist: a new scenario with a contradictory workbook is simply
    not tested, and the suite reports nothing. Walking for `agents.yaml` is the definition of "a
    scenario config folder" the Creator itself uses.
    """
    # `rglob`, not a two-level glob. The Creator treats every subfolder of a config directory as a
    # region with its own `agents.yaml`/`agents.xlsx` (`scenario_run.prepare_config` rglobs for
    # exactly that), so a nested region is a scenario this test has to see. A fixed depth silently
    # skipped one: a planted `scenario_with_grid/region_north/` with a contradictory workbook left
    # the whole suite green, and the same folder one level up failed correctly. Demonstrated.
    found = {path.parent.relative_to(REPO_ROOT).as_posix()
             for root in ('examples', 'tests/e2e/scenarios')
             for path in (REPO_ROOT / root).rglob('agents.yaml')}

    assert found == set(SCENARIOS), (
        f'scenario folders in the tree but not listed: {sorted(found - set(SCENARIOS))}; '
        f'listed but not in the tree: {sorted(set(SCENARIOS) - found)}')


def test_the_scenario_carries_the_keys_this_test_compares(scenario):
    """A comparison over an empty set of values passes and proves nothing.

    So the values are asserted to exist before they are asserted to agree. Renaming
    `ems/controller/rtc/optimization/framework` would otherwise leave every comparison below
    trivially true -- the failure mode #206 was filed about, in a test written to prevent it.
    """
    path, name = scenario

    for key in KEYS:
        assert yaml_values(path / 'agents.yaml', key), f'{name}/agents.yaml states no {key}'

    if name in NO_WORKBOOK:
        assert not (path / 'agents.xlsx').exists(), (
            f'{name} now ships a workbook; remove it from NO_WORKBOOK so it is compared')
        return

    book = path / 'agents.xlsx'
    assert book.exists(), f'{name} ships no agents.xlsx; add it to NO_WORKBOOK if that is intended'
    for key in KEYS:
        per_sheet = workbook_values(book, key)
        assert per_sheet, f'{name}/agents.xlsx has no sheets at all'
        empty = sorted(sheet for sheet, values in per_sheet.items() if not values)
        assert not empty, f'{name}/agents.xlsx sheets {empty} carry no {key} column'


def test_the_workbook_says_what_the_yaml_says(scenario):
    """The fix for #214, read out of both files.

    Every mismatch across every key and every sheet is collected and reported together. Failing on
    the first would hide the rest, and when two files drift it is the *shape* of the drift that
    says whether it was one careless edit or a stale workbook.
    """
    path, name = scenario
    if name in NO_WORKBOOK:
        pytest.skip(f'{name} ships no agents.xlsx; the Creator writes it from the YAML')

    found = mismatches(path / 'agents.yaml', path / 'agents.xlsx', name)
    # Nested regions carry their own pair, and `new_scenario_from_files` reads each one.
    for nested in sorted(path.rglob('agents.xlsx')):
        if nested.parent == path:
            continue
        region = nested.parent
        found += mismatches(region / 'agents.yaml', nested,
                            f'{name}/{region.relative_to(path).as_posix()}')

    assert not found, (
        'a scenario states its backend twice and the two disagree. Whichever file the Creator '
        'entry point reads is what runs, so the other one is a lie to whoever opens it -- see '
        '#214:\n  ' + '\n  '.join(found))


def test_a_sheet_that_carries_no_value_is_a_mismatch_of_its_own(tmp_path):
    """The per-sheet comparison, driven by a workbook no shipped scenario can produce.

    **This test exists because the per-sheet split was vacuous without it.** Collapsing
    `workbook_values` to one set over the whole workbook left every other test in this file green:
    every shipped workbook is internally consistent, so no fixture could tell a per-sheet answer
    from a workbook-wide one. Found by mutation, not by reading.

    The case that separates them is a sheet with **no** value for the key -- a renamed or deleted
    column. A union over the workbook is then `{'poi'}`, exactly what the YAML says, and the
    disagreement is invisible; per sheet, the empty one is a mismatch. That is #206's shape one
    level down and the reason `switch_in_workbook` reports per sheet too.
    """
    from openpyxl import Workbook

    (tmp_path / 'agents.yaml').write_text('        framework: poi\n', encoding='utf-8')
    book = Workbook()
    good = book.active
    good.title = 'ctsp'
    good.append(['', 'ems/controller/rtc/optimization/framework'])
    good.append([0, 'poi'])
    blind = book.create_sheet('industry')
    blind.append(['', 'ems/controller/rtc/optimization/framework_renamed'])
    blind.append([0, 'poi'])
    book.save(tmp_path / 'agents.xlsx')

    found = mismatches(tmp_path / 'agents.yaml', tmp_path / 'agents.xlsx', 'stub',
                       keys=('framework',))

    assert len(found) == 1, found
    assert "sheet 'industry' says []" in found[0], found


def test_a_single_disagreeing_sheet_is_named(tmp_path):
    """And the other direction: one sheet right, one wrong. The report must name the wrong one.

    Without this, "some sheet disagrees" could be reported against the wrong sheet and nobody
    would notice -- the assertion message is the whole product of this test, since a mismatch here
    is something a human has to go and fix in a binary file.
    """
    from openpyxl import Workbook

    (tmp_path / 'agents.yaml').write_text('        framework: poi\n', encoding='utf-8')
    book = Workbook()
    good = book.active
    good.title = 'ctsp'
    good.append(['', 'ems/controller/rtc/optimization/framework'])
    good.append([0, 'poi'])
    wrong = book.create_sheet('industry')
    wrong.append(['', 'ems/controller/rtc/optimization/framework'])
    wrong.append([0, 'linopy'])
    book.save(tmp_path / 'agents.xlsx')

    found = mismatches(tmp_path / 'agents.yaml', tmp_path / 'agents.xlsx', 'stub',
                       keys=('framework',))

    assert len(found) == 1, found
    assert "sheet 'industry' says ['linopy']" in found[0], found
    assert 'ctsp' not in found[0], found


def test_a_yaml_stating_one_key_two_ways_is_caught(tmp_path):
    """The comparison is set-to-set, so a YAML that disagrees with *itself* is a mismatch too.

    Worth its own test because the natural implementation takes the first regex match and would
    silently compare half the file. A scenario whose RTC says `poi` and whose FBC says `linopy`
    runs two backends in one agent, which is exactly the confusion #214 is about.
    """
    (tmp_path / 'agents.yaml').write_text(
        '        framework: poi\n        framework: linopy\n        solver: highs\n',
        encoding='utf-8')

    assert yaml_values(tmp_path / 'agents.yaml', 'framework') == {'poi', 'linopy'}


def test_the_compared_keys_are_the_ones_that_decide_the_backend():
    """`KEYS` is checked against the workbooks, not merely declared.

    Without this it is an allowlist that exempts whatever is left out of it: with `KEYS` reduced to
    `('framework',)` the `scenario_with_market` workbook can be put back into disagreement on
    `solver` — #214, on the file this module exists to repair — and all 15 tests still pass.
    Demonstrated by mutation, not argued.

    The check is that every backend-selecting column present in a shipped workbook is compared.
    `ems/controller/{rtc,fbc}/optimization/` is where the Creator writes them, and its leaf names
    are what `KEYS` holds.
    """
    leaves = set()
    for name in SCENARIOS:
        book = REPO_ROOT / name / 'agents.xlsx'
        if not book.exists():
            continue
        with pd.ExcelFile(book) as workbook:
            for sheet in workbook.sheet_names:
                for column in workbook.parse(sheet, index_col=0).columns:
                    text = str(column)
                    if text.startswith('ems/controller/') and '/optimization/' in text:
                        leaves.add(text.rsplit('/', 1)[-1])

    assert leaves, 'no optimization columns found in any shipped workbook; the layout changed'
    assert leaves == set(KEYS), (
        f'the shipped workbooks carry optimization keys {sorted(leaves)} but this module compares '
        f'{sorted(KEYS)}. A key that is written but not compared is free to disagree with the '
        f'YAML, which is #214')
