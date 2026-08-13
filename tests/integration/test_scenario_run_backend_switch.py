"""`run_example`'s backend switch reaches the file the Creator will actually read (#206).

`run_example(framework=..., solver=...)` edits a scenario's config before running it. Until this
existed it edited `agents.yaml` and nothing else — and a scenario built with
`new_scenario_from_files` gets its agents from **`agents.xlsx`**, which nothing regenerates and
which the YAML has no part in. A caller asking such a scenario for a different backend had the
request accepted, the switch's own `assert switched` satisfied by the YAML it edited, and the
shipped backend run anyway.

Three shipped scenarios were exposed, not one: `grid_golden` and `scenario_with_topology` pin
`poi`/`highs` in their workbooks, and `scenario_with_market` pinned `linopy`/`gurobi` — so asking
that last one for HiGHS would have run **Gurobi**, on a machine that needs a licence for it. (That
workbook now agrees with its own YAML at `poi`/`gurobi`; the disagreement was #214, and
`test_shipped_configs_agree_with_their_workbooks.py` is what keeps it that way. The exposure this
paragraph describes was real and is what the switch had to be fixed for.)

**These tests read the config files back, never the arguments they were given.** That distinction
is the whole point: the defect was a switch that reported success on the strength of what it had
been asked to do. So the workbook is re-opened *the way the Creator opens it* — `pd.ExcelFile`,
`parse(sheet, index_col=0)`, exactly `agents.create_agents_from_file` — and the values are read out
of the parsed frame.

Both halves of the fix are covered here, because both are cheap to cover: the config edit, and
`assert_backend_honoured`, the guard that reads the run's own receipt. The second is the half that
generalises to whatever file a future `creator_method` reads, and it was reachable only through
`run_example` — so until `TestTheReceiptIsChecked` existed, a typo in it failed open until someone
ran a job costing minutes.

What this file cannot see is an actual run honouring an actual config. That is
`tests/e2e/test_ctsp_industry.py::test_the_requested_backend_is_what_solved`, and it is not
redundant with these: it fails on the *consequence* where these fail on the *cause*, so it is the
one that survives someone reinstating a YAML-only switch by a different route. It runs the
`ctsp_industry` fixture rather than `grid_golden` — same assertion, 26-36 s instead of 232-272 s,
and it carries the ctsp/industry coverage with it.
"""
import json
import shutil

import pandas as pd
import pytest
from openpyxl import Workbook

from tests.scenario_run import (REPO_ROOT, assert_backend_honoured, prepare_config,
                                switch_in_workbook)

#: The workbook-built scenario the defect was found against. Its own literals, like every other
#: file that names it (`test_golden_master`, `test_grid_restrictions`, `test_solver_backend_smoke`)
#: — `GoldenScenario` exports no constant to share, so moving the scenario means fixing four files.
CONFIG_ROOT = REPO_ROOT / 'tests' / 'e2e' / 'scenarios'
SCENARIO = 'grid_golden'

#: What `grid_golden` ships. The switch is only meaningful against a value that is not already the
#: one being asked for, so the tests below assert this first rather than assuming it.
SHIPPED = {'framework': 'poi', 'solver': 'highs'}

#: What to ask for instead. Never solved here — nothing in this file runs a model — so `gurobi`
#: needs no licence and is the better choice precisely because no environment defaults to it.
REQUESTED = {'framework': 'linopy', 'solver': 'gurobi'}


def workbook_values(path, key):
    """Every value of every `.../<key>` column, read the way `create_agents_from_file` reads it.

    Closed explicitly: pandas defines no `__del__`, so an unclosed `ExcelFile` keeps a Windows file
    handle open until the next gc pass, and `tmp_path` cleanup then fails with `WinError 32`.
    """
    values = []
    with pd.ExcelFile(path) as book:
        for sheet in book.sheet_names:
            frame = book.parse(sheet, index_col=0)
            for column in frame.columns:
                if str(column).rsplit('/', 1)[-1] == key:
                    values.extend(frame[column].dropna().tolist())
    return values


@pytest.fixture(scope='module')
def shipped_workbook():
    """The unmodified workbook, so "the switch changed it" is a comparison and not an assumption."""
    return {key: workbook_values(CONFIG_ROOT / SCENARIO / 'agents.xlsx', key) for key in SHIPPED}


def test_the_scenario_still_ships_a_backend_worth_switching_away_from(shipped_workbook):
    """The premise. If `grid_golden` ever ships `linopy`, the tests below stop proving anything.

    Stated separately so that a fixture change fails *here*, naming the cause, rather than turning
    the switch tests green for the wrong reason — which is the exact failure mode #206 is about.
    """
    yaml_text = (CONFIG_ROOT / SCENARIO / 'agents.yaml').read_text(encoding='utf-8')

    for key, shipped in SHIPPED.items():
        assert shipped_workbook[key], f'{SCENARIO}/agents.xlsx has no {key} column at all'
        assert set(shipped_workbook[key]) == {shipped}, (
            f'{SCENARIO}/agents.xlsx no longer ships {key}={shipped!r} throughout; it holds '
            f'{sorted(set(shipped_workbook[key]))}. Re-check what REQUESTED should be')
        assert shipped != REQUESTED[key], f'asking for the shipped {key} proves nothing'
        # The YAML too, and for its own reason: `test_the_switch_still_reaches_the_yaml` compares
        # against `SHIPPED`, so a fixture that started shipping `linopy` there would leave that
        # test passing over a switch that did nothing. Checking only the workbook would be #206
        # one file over -- a premise guard that guards the file it was written for and no other.
        assert f'{key}: {shipped}' in yaml_text, (
            f'{SCENARIO}/agents.yaml no longer ships {key}: {shipped}, so asserting the switch '
            f'moved it away from that value proves nothing')


def test_the_switch_reaches_the_workbook_the_scenario_is_built_from(tmp_path, shipped_workbook):
    """The fix for #206, read back out of the file rather than off the request.

    Goes through `prepare_config` — the code path `run_example` itself takes — rather than calling
    `switch_in_workbook` directly, so removing the workbook step from `run_example` fails this test
    instead of leaving it passing against a helper nothing calls any more.
    """
    config = prepare_config(tmp_path, CONFIG_ROOT, SCENARIO, framework=REQUESTED['framework'],
                            solver=REQUESTED['solver'])

    for key, requested in REQUESTED.items():
        values = workbook_values(config / 'agents.xlsx', key)
        assert values, f'the switch emptied the {key} column'
        assert set(values) == {requested}, (
            f'agents.xlsx still holds {key}={sorted(set(values))} after asking for {requested!r}, '
            f'so a scenario built with new_scenario_from_files would run the shipped backend')
        # Same number of cells as before: a switch replaces values, it does not add or drop agents.
        assert len(values) == len(shipped_workbook[key])


def test_the_switch_still_reaches_the_yaml(tmp_path):
    """The workbook must not have replaced the YAML edit. Both files carry the key; both move.

    `grid_golden/agents.yaml` is inert for this scenario, but a config file that looks
    authoritative, is not, and disagrees with the file that is, is how the next version of #206
    gets written.
    """
    config = prepare_config(tmp_path, CONFIG_ROOT, SCENARIO, framework=REQUESTED['framework'],
                            solver=REQUESTED['solver'])
    text = (config / 'agents.yaml').read_text(encoding='utf-8')

    for key, requested in REQUESTED.items():
        assert f'{key}: {SHIPPED[key]}' not in text, f'agents.yaml still names the shipped {key}'
        assert f'{key}: {requested}' in text


def test_the_shipped_scenario_itself_is_never_modified(tmp_path, shipped_workbook):
    """`prepare_config` copies first. A switch that edited the repository's own config would leave
    every later test in the session running a backend nobody asked for."""
    prepare_config(tmp_path, CONFIG_ROOT, SCENARIO, framework=REQUESTED['framework'],
                   solver=REQUESTED['solver'])

    for key in SHIPPED:
        assert workbook_values(CONFIG_ROOT / SCENARIO / 'agents.xlsx', key) == shipped_workbook[key]


def test_a_workbook_with_no_such_column_reports_nothing_switched(tmp_path):
    """`switch_in_workbook` reports 0 rather than pretending. The caller's half is two tests down."""
    path = tmp_path / 'agents.xlsx'
    book = Workbook()
    book.active.append(['', 'general/agent_id', 'ems/controller/rtc/optimization/backend'])
    book.active.append([0, 'agent_0', 'poi'])
    book.save(path)

    assert switch_in_workbook(path, 'framework', 'linopy') == {'Sheet': 0}


def test_a_sheet_that_did_not_switch_is_reported_separately(tmp_path):
    """One sheet must not vouch for another — #206's shape one level down.

    `create_agents_file_from_config` writes one sheet per agent type, and `config_templates` ships
    five. A workbook-wide count of "1 cell switched" is satisfied by the first sheet while every
    other agent type silently keeps the shipped backend, which is exactly the class of failure this
    whole change exists to close.
    """
    path = tmp_path / 'agents.xlsx'
    book = Workbook()
    sfh = book.active
    sfh.title = 'sfh'
    sfh.append(['', 'ems/controller/rtc/optimization/framework'])
    sfh.append([0, 'poi'])
    industry = book.create_sheet('industry')
    industry.append(['', 'ems/controller/rtc/optimization/framework_renamed'])
    industry.append([0, 'poi'])
    book.save(path)

    switched = switch_in_workbook(path, 'framework', 'linopy')

    assert switched == {'sfh': 1, 'industry': 0}, (
        'a per-sheet report is what lets the caller see that "industry" kept its backend; a total '
        'of 1 would have read as success')


def test_a_workbook_containing_formulas_is_refused(tmp_path):
    """Saving through openpyxl blanks cached formula results, which the Creator reads as unset.

    So a *backend* switch would silently empty an unrelated sizing column. Refusing is the only
    honest option: the alternative is a config edit that quietly changes the scenario. Verified
    against the four shipped workbooks — none contains a formula, so this refuses nothing real.
    """
    path = tmp_path / 'agents.xlsx'
    book = Workbook()
    book.active.append(['', 'ems/controller/rtc/optimization/framework', 'general/parameters/area'])
    book.active.append([0, 'poi', '=10*15'])
    book.save(path)

    with pytest.raises(AssertionError, match='contains formulas'):
        switch_in_workbook(path, 'framework', 'linopy')


def test_a_workbook_that_cannot_be_switched_fails_the_run_loudly(tmp_path):
    """The fail-loud half, through `prepare_config` — the assert, not the return value it reads.

    A renamed column is exactly what the old switch sailed through, and `switch_in_workbook`
    returning 0 only helps if someone acts on it. Testing the helper alone left that assert
    uncovered: deleting it kept every other test in this file green, which is the shape of #206
    again — a check that reports success because nothing consults the answer.
    """
    config_root = tmp_path / 'configs'
    scenario = config_root / SCENARIO
    shutil.copytree(CONFIG_ROOT / SCENARIO, scenario)

    from openpyxl import load_workbook
    book = load_workbook(scenario / 'agents.xlsx')
    for sheet in book.worksheets:
        for (cell,) in sheet.iter_cols(min_row=1, max_row=1):
            if cell.value is not None and str(cell.value).endswith('/framework'):
                cell.value = f'{cell.value}_renamed'
    book.save(scenario / 'agents.xlsx')

    with pytest.raises(AssertionError, match='no framework value to switch'):
        prepare_config(tmp_path / 'run', config_root, SCENARIO, framework=REQUESTED['framework'])


def test_one_sheet_switching_does_not_vouch_for_the_others(tmp_path):
    """The caller's half of the per-sheet report, through `prepare_config`.

    `grid_golden` ships a single `sfh` sheet, so no shipped fixture can exercise this — a second
    agent type is added here. Without it the caller could go back to asserting a workbook-wide
    total and every test would stay green, which is how the original defect survived review.
    """
    config_root = tmp_path / 'configs'
    scenario = config_root / SCENARIO
    shutil.copytree(CONFIG_ROOT / SCENARIO, scenario)

    from openpyxl import load_workbook
    book = load_workbook(scenario / 'agents.xlsx')
    industry = book.create_sheet('industry')
    industry.append(['', 'ems/controller/rtc/optimization/framework_renamed'])
    industry.append([0, 'poi'])
    book.save(scenario / 'agents.xlsx')

    with pytest.raises(AssertionError, match=r"sheets \['industry'\]"):
        prepare_config(tmp_path / 'run', config_root, SCENARIO, framework=REQUESTED['framework'])


def test_a_blank_cell_stays_blank(tmp_path):
    """An agent with the controller switched off has no backend, and must not acquire one.

    Writing `linopy` into a blank would switch a controller *on* in the name of switching a
    backend, which changes the scenario rather than the backend it runs under.
    """
    path = tmp_path / 'agents.xlsx'
    book = Workbook()
    book.active.append(['', 'ems/controller/rtc/optimization/framework'])
    book.active.append([0, 'poi'])
    book.active.append([1, None])
    book.save(path)

    assert switch_in_workbook(path, 'framework', 'linopy') == {'Sheet': 1}
    assert workbook_values(path, 'framework') == ['linopy']


def test_an_unrelated_column_named_like_the_key_is_left_alone(tmp_path):
    """The key is the last path segment, so `.../framework` matches and `.../framework_note`
    does not. Matching on a substring would rewrite whatever happened to contain the word."""
    path = tmp_path / 'agents.xlsx'
    book = Workbook()
    book.active.append(['', 'ems/controller/rtc/optimization/framework',
                        'ems/controller/rtc/optimization/framework_note'])
    book.active.append([0, 'poi', 'poi'])
    book.save(path)

    assert switch_in_workbook(path, 'framework', 'linopy') == {'Sheet': 1}
    with pd.ExcelFile(path) as book:
        frame = book.parse(0, index_col=0)
    assert frame['ems/controller/rtc/optimization/framework_note'].tolist() == ['poi']


class TestTheReceiptIsChecked:
    """`assert_backend_honoured` — the half of the fix that generalises, covered where it is cheap.

    Switching the workbook fixes the one file that was missed; this reads the record the run wrote
    and refuses to accept a run that used something else, whatever file a future `creator_method`
    consults. It is the more important half and it was reachable only through `run_example`, which
    means only from `tests/e2e/` — so a typo in it would have failed open for minutes at a time,
    or until someone ran the slow jobs. It is a pure function over a JSON path; there is no reason
    for its coverage to cost a simulation.
    """

    @staticmethod
    def record(tmp_path, pairs):
        path = tmp_path / 'backends.json'
        path.write_text(json.dumps(pairs), encoding='utf-8')
        return path

    def test_a_matching_record_is_accepted(self, tmp_path):
        assert_backend_honoured(self.record(tmp_path, [['linopy', 'highs']]), 'linopy', 'highs')

    def test_a_different_framework_is_refused(self, tmp_path):
        with pytest.raises(AssertionError, match=r"asked for framework 'linopy'"):
            assert_backend_honoured(self.record(tmp_path, [['poi', 'highs']]), 'linopy', 'highs')

    def test_a_different_solver_is_refused(self, tmp_path):
        with pytest.raises(AssertionError, match=r"asked for solver 'highs'"):
            assert_backend_honoured(self.record(tmp_path, [['poi', 'gurobi']]), None, 'highs')

    def test_a_run_that_used_two_backends_is_refused(self, tmp_path):
        """Half the agents honouring the switch is not the switch being honoured."""
        pairs = [['linopy', 'highs'], ['poi', 'highs']]
        with pytest.raises(AssertionError, match=r'actually used \[.linopy., .poi.\]'):
            assert_backend_honoured(self.record(tmp_path, pairs), 'linopy', 'highs')

    def test_a_run_that_solved_nothing_is_refused(self, tmp_path):
        """The failure mode an empty-collection check invites: nothing solved, so nothing to
        disagree with, so a naive `all(...)` over the record would pass."""
        with pytest.raises(AssertionError, match='without building or solving a single model'):
            assert_backend_honoured(self.record(tmp_path, []), 'linopy', 'highs')

    def test_a_missing_record_is_refused(self, tmp_path):
        with pytest.raises(AssertionError, match='wrote no backend record'):
            assert_backend_honoured(tmp_path / 'absent.json', 'linopy', 'highs')

    def test_an_axis_that_was_not_asked_for_is_not_checked(self, tmp_path):
        """`framework='linopy'` with no `solver` is a legitimate call — `test_backend_equivalence`
        makes it — and must not start demanding that the caller name a solver as well."""
        assert_backend_honoured(self.record(tmp_path, [['linopy', 'gurobi']]), 'linopy', None)


def test_a_scenario_without_a_workbook_is_unaffected(tmp_path):
    """`simple_scenario` ships no `agents.xlsx` — the Creator writes one from the YAML — so the
    workbook step must be skipped rather than fail on a missing file."""
    example = REPO_ROOT / 'examples' / 'create_simple_scenario'
    assert not (example / 'simple_scenario' / 'agents.xlsx').exists(), (
        'simple_scenario now ships a workbook; this test no longer covers the no-workbook case')

    config = prepare_config(tmp_path, example, 'simple_scenario', framework='linopy')

    assert 'framework: linopy' in (config / 'agents.yaml').read_text(encoding='utf-8')
