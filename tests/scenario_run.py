"""Running a shipped example and reducing its results to a comparable fingerprint.

Extracted from `tests/e2e/test_golden_master.py` so that the golden master and the backend
equivalence tests reduce results the *same* way. Two different fingerprints would make the two
tests incomparable, and the whole value of the equivalence test is that its linopy arm reproduces
the golden reference exactly.

Reproducibility rests on seeding `random` and `numpy.random` and pinning `PYTHONHASHSEED`: the
Creator draws agent ids, plant ownership and sizings from all three.
"""
import json
import os
import re
import shutil
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
SEED = 20260804

# Runs Creator then Executor in a subprocess. A subprocess rather than an in-process call because
# `PYTHONHASHSEED` only takes effect at interpreter start, and the Creator's file selection depends
# on it.
RUNNER = """
import os, random
import numpy as np
random.seed({seed})
np.random.seed({seed})
from hamlet import Creator, Executor
{probe}
Creator(path=r"{config_dir}").{creator_method}()
Executor(r"{scenarios}/{name}", num_workers=1).run()
print("RUN_OK")
"""

# Records every (framework, solver) pair that actually built or solved a model during the run, so
# a caller can assert the run used what it asked for instead of trusting that a config edit took.
#
# This is not defensiveness. In !212 the `framework:` switch below matched a literal shipped value
# and silently became a no-op when the default flipped, which would have made both arms of a
# backend comparison run the same backend and agree with each other. A config edit is a *request*;
# this is the receipt. `num_workers=1` is what makes it work: every solve happens in this process.
#
# The two POI modules are patched by name rather than `poi_solver.create_model`, because each does
# `from ... import create_model`, which binds the function into its own namespace -- rebinding it
# on `poi_solver` afterwards would patch nothing, record an empty set, and fail open.
BACKEND_PROBE = '''
import atexit, json
import linopy
from hamlet.executor.utilities.controller.fbc.mpc.poi import mpc_poi
from hamlet.executor.utilities.controller.rtc.optim.poi import optim_poi

_used = set()

for _module in (mpc_poi, optim_poi):
    _original_create = _module.create_model

    def _create_model(solver, _original_create=_original_create):
        _used.add(('poi', solver))
        return _original_create(solver)

    _module.create_model = _create_model

_original_solve = linopy.Model.solve


def _solve(self, *args, **kwargs):
    _used.add(('linopy', kwargs.get('solver_name') or (args[0] if args else None)))
    return _original_solve(self, *args, **kwargs)


linopy.Model.solve = _solve

atexit.register(lambda: open(r"{record}", "w", encoding="utf-8").write(json.dumps(sorted(_used))))
'''


def switch_in_yaml(text, key, value):
    """Point every `key:` in a YAML config at `value`, or fail saying there was none to point."""
    text, switched = re.subn(rf'^(\s*){key}: *\w+', rf'\g<1>{key}: {value}', text,
                             flags=re.MULTILINE)
    assert switched, f'no {key} key to switch in agents.yaml'
    return text


def switch_in_workbook(path, key, value):
    """Point every `.../<key>` column of every sheet in an agents workbook at `value`.

    The workbook stores what the YAML nests: `ems/controller/rtc/optimization/framework` is one
    column header, so the key is its last path segment.

    Edited in place with openpyxl rather than round-tripped through pandas, so a switch touches the
    targeted cells and nothing else. Only cells that already hold a value are rewritten: a blank
    means the agent has no such controller, and filling one in would be adding configuration rather
    than switching it.

    **Returns cells rewritten per sheet, not one total.** A single number lets one matching sheet
    vouch for the rest -- `create_agents_file_from_config` writes one sheet per agent type
    (`agents.py:228`), and `config_templates/agents.yaml` declares five -- so a workbook whose
    `sfh` sheet switches and whose `industry` sheet has a renamed column would report success while
    half its agents kept the shipped backend. That is #206's own shape one level down, and a total
    cannot see it.
    """
    from openpyxl import load_workbook

    book = load_workbook(path)

    # openpyxl writes no cached formula results and pandas reads only cached results, so re-saving
    # this workbook blanks every formula in it -- and the Creator reads a blank as "not configured".
    # A backend switch silently rewriting an unrelated sizing column is worse than not switching, so
    # this refuses rather than proceeds. None of the four shipped workbooks contains a formula.
    formulas = [f'{sheet.title}!{cell.coordinate}'
                for sheet in book.worksheets for row in sheet.iter_rows() for cell in row
                if isinstance(cell.value, str) and cell.value.startswith('=')]
    assert not formulas, (
        f'{path.name} contains formulas ({formulas[:5]}), and saving it through openpyxl would '
        f'blank their cached values -- which the Creator would then read as unset. Switch the '
        f'backend in this scenario some other way, or replace the formulas with their values')

    switched = {}
    for sheet in book.worksheets:
        headers = [cell for (cell,) in sheet.iter_cols(min_row=1, max_row=1)
                   if cell.value is not None and str(cell.value).rsplit('/', 1)[-1] == key]
        count = 0
        for header in headers:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=header.column)
                if cell.value in (None, ''):
                    continue
                cell.value = value
                count += 1
        switched[sheet.title] = count

    if any(switched.values()):
        book.save(path)
    return switched


def assert_backend_honoured(record, framework, solver):
    """The request was a request; this reads the receipt the run wrote and checks it was honoured.

    **This is the general guard, and it is the half of #206 that generalises.** The switch reaching
    `agents.xlsx` fixes the one file that was missed; this fixes the class, because it does not care
    which file a `creator_method` reads -- only that the run reports having used what was asked for.

    Deliberately checks each axis only when it was asked for. `framework='linopy'` with no `solver`
    is a legitimate call -- `test_backend_equivalence` makes it -- and must not start requiring the
    caller to name a solver as well.
    """
    assert record.exists(), (
        'the run completed but wrote no backend record, so what solved it is unknown')
    used = {tuple(pair) for pair in json.loads(record.read_text(encoding='utf-8'))}

    assert used, (
        f'the run completed without building or solving a single model, so the requested backend '
        f'({framework or "any"} + {solver or "any"}) cannot have been used; the request was lost')

    for axis, requested, actual in (('framework', framework, sorted({pair[0] for pair in used})),
                                    ('solver', solver, sorted({pair[1] for pair in used}))):
        if requested is None:
            continue
        assert actual == [requested], (
            f'asked for {axis} {requested!r}, but the run actually used {actual} '
            f'(pairs: {sorted(used)}). The config edit did not reach the file this scenario is '
            f'built from -- see #206')


def table_kind(path, root):
    """The table's identity, with the random agent id replaced by its type."""
    parts = Path(path).relative_to(root).parts
    # agents/<type>/<random id>/<table>.ft  ->  agents/<type>/<table>.ft
    if parts and parts[0] == 'agents' and len(parts) == 4:
        return f'agents/{parts[1]}/{parts[3]}'
    return '/'.join(parts)


def fingerprint(results_root):
    """Row counts and per-column statistics, grouped by table kind.

    Both result formats are read. Agent and market tables are Feather; the grid stage writes its
    power flow results as CSV (`res_bus`, `res_trafo`, `res_line`, ...), so globbing `*.ft` alone
    silently pinned nothing the grid produces -- which for a scenario added specifically to cover
    the grid stage would have been a golden master that passed while every number in it moved.
    Scenarios with `electricity.active: False` write no CSV at all and are unaffected.
    """
    import polars as pl

    grouped = defaultdict(lambda: {'files': 0, 'rows': 0, 'columns': {}})
    paths = sorted(list(Path(results_root).rglob('*.ft')) + list(Path(results_root).rglob('*.csv')))
    for path in paths:
        frame = (pl.read_csv(path) if path.suffix == '.csv'
                 else pl.read_ipc(path, memory_map=False))
        entry = grouped[table_kind(path, results_root)]
        entry['files'] += 1
        entry['rows'] += len(frame)
        for column, dtype in zip(frame.columns, frame.dtypes):
            if not dtype.is_numeric() or frame[column].null_count() == len(frame):
                continue
            stats = entry['columns'].setdefault(column, {'sum': 0.0, 'min': None, 'max': None})
            stats['sum'] += float(frame[column].sum() or 0)
            low, high = frame[column].min(), frame[column].max()
            if low is not None:
                stats['min'] = float(low) if stats['min'] is None else min(stats['min'], float(low))
                stats['max'] = float(high) if stats['max'] is None else max(stats['max'], float(high))

    return {kind: entry for kind, entry in sorted(grouped.items())}


def prepare_config(base, example_dir, scenario_name, framework=None, solver=None, edits=(),
                   config_edits=None):
    """Copy an example's config under `base` and apply every requested edit to it.

    Split out of `run_example` so the edits can be checked without paying for a run: the backend
    switch has to reach the file the Creator reads, and #206 went *unnoticed* for as long as
    observing that meant running a whole scenario and reading the results afterwards.

    Returns the config directory. See `run_example` for what each argument means.
    """
    scenarios, results = base / 'scenarios', base / 'results'
    config = base / scenario_name
    shutil.copytree(example_dir / scenario_name, config)
    scenarios.mkdir(exist_ok=True)
    results.mkdir(exist_ok=True)

    setup = config / 'setup.yaml'
    text = setup.read_text(encoding='utf-8')
    for old, new in (('input: ../../input_data', f'input: {(REPO_ROOT / "input_data").as_posix()}'),
                     ('scenarios: ../../scenarios', f'scenarios: {scenarios.as_posix()}'),
                     ('results: ../../results', f'results: {results.as_posix()}')):
        assert old in text, f'{old!r} not found in setup.yaml'
        text = text.replace(old, new)
    setup.write_text(text, encoding='utf-8')

    agents = config / 'agents.yaml'
    agents_text = agents.read_text(encoding='utf-8')
    if framework is not None:
        agents_text = switch_in_yaml(agents_text, 'framework', framework)
    if solver is not None:
        agents_text = switch_in_yaml(agents_text, 'solver', solver)
    for old, new in edits:
        assert old in agents_text, f'{old!r} not found in agents.yaml'
        agents_text = agents_text.replace(old, new)
    agents.write_text(agents_text, encoding='utf-8')

    # The workbooks, where the scenario ships any. `new_scenario_from_files` builds the agents from
    # them and nothing regenerates them, so this is the *only* place their backend can be switched;
    # the other two entry points rewrite them (`__create_agent_files` passes `overwrite=True`) from
    # the `ems` block `fill_ems` reads out of the YAML above, which makes this edit redundant there
    # rather than wrong. Either way both files end up saying the same thing. #206.
    #
    # `rglob` rather than `config / 'agents.xlsx'`: the Creator treats every subfolder of the config
    # directory as a region and creates agents for each (`setup.py.__loop_through_dict`), so a
    # nested scenario has one workbook per region. No shipped scenario is nested, so today this
    # finds exactly one file -- but editing only the root would be this same defect again, and the
    # YAML edit above still has that gap (the run-time receipt is what covers it).
    for workbook in sorted(config.rglob('agents.xlsx')):
        for key, value in (('framework', framework), ('solver', solver)):
            if value is None:
                continue
            switched = switch_in_workbook(workbook, key, value)
            missed = sorted(sheet for sheet, count in switched.items() if not count)
            assert switched and not missed, (
                f'no {key} value to switch in {workbook.relative_to(config.parent)}'
                + (f', sheets {missed}' if missed else '')
                + f', so the request for {value!r} would be silently lost for those agents if this '
                  f'scenario is built with new_scenario_from_files')

    for filename, replacements in (config_edits or {}).items():
        target = config / filename
        assert target.exists(), f'{filename} not found in {scenario_name}'
        content = target.read_text(encoding='utf-8')
        for old, new in replacements:
            assert old in content, f'{old!r} not found in {filename}'
            content = content.replace(old, new)
        target.write_text(content, encoding='utf-8')

    return config


def run_example(base, example_dir, scenario_name, framework=None, solver=None, edits=(),
                config_edits=None, record_backends=None,
                creator_method='new_scenario_from_configs'):
    """Run one example end to end under `base`, and return the fingerprint of its results.

    `creator_method` is the Creator entry point the example's own notebook calls — the three
    shipped ones differ (`new_scenario_from_configs` reads the YAML, `new_scenario_from_files`
    reads `agents.xlsx`, `new_scenario_from_grids` derives the agents from the grid file), and
    which one is used decides whether the agent ids are drawn fresh or come from a file. That
    matters for the `topology` grid method, whose topology file names agents by id.

    `framework` switches every `framework:` key to the named backend and `solver` every `solver:`
    key to the named solver, whatever the config ships; None leaves either alone. Both match the
    *key* rather than the shipped value on purpose -- the framework switch used to look for the
    literal `framework: linopy`, which silently became a no-op the moment the default flipped to
    `poi`, and a no-op here means both arms of a backend comparison run the same backend and
    agree. `edits` is a sequence of (old, new) string replacements applied to `agents.yaml`, and
    `config_edits` is a {filename: [(old, new), ...]} mapping for any other config file --
    `grids.yaml`, say, to switch a grid restriction on.

    **The backend switch reaches the workbook as well as the YAML** -- see `prepare_config` for
    which file each entry point actually consults, and for what it still does not reach. Both are
    switched where both exist, so neither can be the authoritative one while the other quietly
    disagrees with it (#206).

    Every replacement must match at least once, in every file it is applied to. That is deliberate:
    a renamed config key then fails the test loudly, instead of the test quietly running an
    unmodified scenario and passing for the wrong reason.

    `record_backends` is a path to write the JSON list of (framework, solver) pairs the run
    actually used. **Whenever `framework` or `solver` is passed the record is written and checked
    here regardless**; passing `record_backends` additionally hands it to the caller. See
    `assert_backend_honoured`.
    """
    scenarios, results = base / 'scenarios', base / 'results'
    config = prepare_config(base, example_dir, scenario_name, framework=framework, solver=solver,
                            edits=edits, config_edits=config_edits)

    # A backend was asked for, so the run has to say what it used -- whether or not the caller
    # wants the record for itself. The alternative is the state #206 was filed about: the receipt
    # existed, was opt-in, and neither the golden master nor the grid tests took it.
    requested = framework is not None or solver is not None
    record = Path(record_backends) if record_backends is not None else base / 'backends_used.json'
    wanted = requested or record_backends is not None
    probe = BACKEND_PROBE.format(record=record.as_posix()) if wanted else ''

    script = RUNNER.format(config_dir=config.as_posix(), seed=SEED, scenarios=scenarios,
                           name=scenario_name, probe=probe, creator_method=creator_method)
    completed = subprocess.run(
        [sys.executable, '-c', script], capture_output=True, text=True,
        encoding='utf-8', errors='replace', timeout=3600,
        env={**os.environ, 'MPLBACKEND': 'Agg', 'PYTHONIOENCODING': 'utf-8',
             'PYTHONHASHSEED': '0'})
    assert 'RUN_OK' in completed.stdout, completed.stderr[-4000:]

    if requested:
        assert_backend_honoured(record, framework, solver)

    return fingerprint(results / scenario_name)


def compare(a, b, label_a, label_b, tolerance):
    """Differences between two fingerprints, as a list of human-readable strings.

    Empty means the two runs agree to `tolerance`. Reported all at once rather than failing on the
    first, because when a model changes it is the shape of the difference across tables that says
    whether it was intended.
    """
    differences = []

    only_a, only_b = sorted(set(a) - set(b)), sorted(set(b) - set(a))
    if only_a:
        differences.append(f'tables only in {label_a}: {only_a}')
    if only_b:
        differences.append(f'tables only in {label_b}: {only_b}')

    for kind in sorted(set(a) & set(b)):
        if a[kind]['rows'] != b[kind]['rows']:
            differences.append(
                f"{kind}: {a[kind]['rows']} rows in {label_a}, {b[kind]['rows']} in {label_b}")
        for column, stats in a[kind]['columns'].items():
            reference = b[kind]['columns'].get(column)
            if reference is None:
                differences.append(f'{kind}:{column} is present in {label_a} only')
                continue
            for statistic, value in stats.items():
                other = reference.get(statistic)
                if value is None or other is None:
                    continue
                scale = max(abs(value), abs(other))
                if scale < 1e-9 or abs(value - other) / scale <= tolerance:
                    continue
                differences.append(
                    f'{kind}:{column}.{statistic}  {value:.6g} ({label_a}) '
                    f'vs {other:.6g} ({label_b})')

    return differences
